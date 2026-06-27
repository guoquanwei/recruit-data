#!/usr/bin/env python3
import argparse
import calendar
import json
import math
import re
from html import escape
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


CHANNEL_ORDER = ["回流", "内推", "渠道社招", "渠道校招", "自主社招"]
TARGET_INTERVIEW_PASS_RATE = 0.7
TARGET_PASS_TO_TRAIN_RATE = 0.4
TARGET_INTERVIEW_TO_TRAIN_RATE = TARGET_INTERVIEW_PASS_RATE * TARGET_PASS_TO_TRAIN_RATE
PASS_FEEDBACK_RESULTS = {"推荐", "强烈推荐"}


def parse_int(value):
    if value is None or value == "":
        return 0
    return int(float(value))


def parse_percent(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return float(str(value).replace("%", "").strip())


def format_percent(value):
    return f"{value:.2f}%"


def safe_text(value):
    return "" if value is None else str(value)


def recruiter_name(value):
    text = safe_text(value).strip()
    if "+" in text:
        return text.split("+", 1)[0].strip()
    return text


def format_date(value):
    if value is None or value == "" or pd.isna(value):
        return ""
    return pd.Timestamp(value).strftime("%Y-%m-%d")


def parse_department(row, table_type):
    if table_type == "active":
        dept = safe_text(row.get("部门", "")).strip()
        sep = "/"
    else:
        dept = safe_text(row.get("离职前部门", "")).strip()
        sep = "-"

    office = safe_text(row.get("办公地点", "")).strip()
    parts = [part.strip() for part in dept.split(sep) if part and part.strip()]
    ops_idx = next((idx for idx, part in enumerate(parts) if part.upper() == "OPS"), None)
    after = parts[ops_idx + 1 :] if ops_idx is not None else parts
    l1 = after[0] if len(after) > 0 else ""
    l2 = after[1] if len(after) > 1 else ""
    l3 = after[2] if len(after) > 2 else ""

    if l1 in {"京津冀大区", "长春基地"} and l2:
        return l2
    if l1 == "江苏基地" and l3:
        return f"{l1}-{l3[:2]}"
    if l1 == "湖南基地":
        if "荷花" in office or "荷花" in dept:
            return "湖南基地-荷花"
        if "空港" in office or "空港" in dept:
            return "湖南基地-空港"
        return "湖南基地"
    if l1 == "济南基地":
        if "济阳" in office or "济阳" in dept:
            return "济南基地-济阳"
        if "夏都" in office or "夏都" in dept:
            return "济南基地-夏都"
        return "济南基地"
    return l1


def read_employee_detail(path, table_type, start_date, end_date):
    if path is None:
        return []
    df = pd.read_excel(path, dtype=object)
    df = df.apply(lambda col: col.map(lambda value: value.strip() if isinstance(value, str) else value))
    df["入培时间_dt"] = pd.to_datetime(df.get("入培时间"), errors="coerce")
    df["入职日期_dt"] = pd.to_datetime(df.get("入职日期"), errors="coerce")
    df["离职日期_dt"] = pd.to_datetime(df.get("离职日期"), errors="coerce")
    df["所属基地"] = df.apply(lambda row: parse_department(row, table_type), axis=1)
    scoped = df[(df["入培时间_dt"] >= start_date) & (df["入培时间_dt"] <= end_date)].copy()
    scoped["工号"] = scoped.get("工号", "").map(safe_text)
    scoped = scoped[scoped["工号"] != ""]
    scoped = scoped.drop_duplicates(subset=["工号"], keep="last")

    records = []
    for _, row in scoped.iterrows():
        hire_date = row.get("入职日期_dt")
        leave_date = row.get("离职日期_dt")
        day_end = leave_date if pd.notna(leave_date) else end_date
        days = 0
        if pd.notna(hire_date) and pd.notna(day_end):
            days = int((day_end - hire_date).days) + 1
        records.append(
            {
                "所属基地": safe_text(row.get("所属基地")),
                "工号": safe_text(row.get("工号")),
                "姓名": safe_text(row.get("姓名")),
                "手机号": safe_text(row.get("手机号码")),
                "招聘渠道": safe_text(row.get("招聘渠道")),
                "渠道名称": safe_text(row.get("渠道名称")),
                "入培日期": format_date(row.get("入培时间_dt")),
                "在职状态": safe_text(row.get("员工状态")) if table_type == "active" else "离职",
                "离职日期": format_date(leave_date),
                "在职天数": max(days, 0),
            }
        )
    return records


def build_detail_index(active_path, leave_path, year, month):
    start_date = pd.Timestamp(year, month, 1)
    end_date = pd.Timestamp(year, month, calendar.monthrange(year, month)[1])
    records = read_employee_detail(active_path, "active", start_date, end_date)
    records.extend(read_employee_detail(leave_path, "leave", start_date, end_date))
    channel_index = {}
    recruiter_index = {}
    for record in records:
        key = (record["所属基地"], record["招聘渠道"])
        channel_index.setdefault(key, []).append(record)
        if record["招聘渠道"] == "自主社招":
            name = recruiter_name(record["渠道名称"])
            if name:
                recruiter_index.setdefault(name, []).append(record)
    for rows in list(channel_index.values()) + list(recruiter_index.values()):
        rows.sort(key=lambda item: (item["入培日期"], item["工号"]))
    return channel_index, recruiter_index


def read_progress_rows(workbook):
    ws = workbook["达成进度"]
    headers = [cell.value for cell in ws[1]]
    rows = []
    current_base = ""
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if row.get("基地"):
            current_base = row["基地"]
        row["基地"] = current_base
        if not row.get("基地") or not row.get("渠道"):
            continue
        row["月度目标"] = parse_int(row.get("月度目标"))
        row["截止目标"] = parse_int(row.get("截止目标"))
        row["实际入培数"] = parse_int(row.get("实际入培数"))
        row["GAP"] = parse_int(row.get("GAP"))
        rows.append(row)
    return rows


def read_efficiency(workbook):
    ws = workbook["人效核算"]
    headers = [cell.value for cell in ws[1]]
    details = []
    summary = []
    summary_header = None

    for values in ws.iter_rows(min_row=2, values_only=True):
        first = values[0]
        if first == "统计维度":
            summary_header = [value for value in values if value is not None]
            continue
        if summary_header:
            if first:
                summary.append(dict(zip(summary_header, values[: len(summary_header)])))
            continue
        if not first:
            continue
        detail = dict(zip(headers, values))
        details.append(detail)

    for row in details:
        row["截止参培达成"] = parse_int(row.get("截止参培达成"))
        row["截止7天参培达成"] = parse_int(row.get("截止7天参培达成"))
        row["截止参培达成率_value"] = parse_percent(row.get("截止参培达成率"))
        row["截止7天参培达成率_value"] = parse_percent(row.get("截止7天参培达成率"))

    for row in summary:
        row["招聘规模"] = parse_int(row.get("招聘规模"))
        row["截止参培达成"] = parse_int(row.get("截止参培达成"))
        row["截止7天参培达成"] = parse_int(row.get("截止7天参培达成"))

    details.sort(key=lambda item: (item["截止参培达成率_value"], safe_text(item.get("姓名"))))
    return summary, details


def build_base_risks(progress_rows):
    totals = [
        row
        for row in progress_rows
        if row["基地"] != "整体"
        and row["渠道"] == "合计"
        and not (row["月度目标"] == 0 and row["实际入培数"] == 0)
    ]
    for row in totals:
        row["风险等级"] = "未达成" if row["GAP"] < 0 else "已达成"
    return sorted(totals, key=lambda row: (0 if row["GAP"] < 0 else 1, row["GAP"], row["基地"]))


def channel_sort_key(row):
    order = CHANNEL_ORDER.index(row["渠道"]) if row["渠道"] in CHANNEL_ORDER else len(CHANNEL_ORDER)
    return row["GAP"], order


def build_unmet_reasons(progress_rows, base_risks, detail_index=None):
    detail_index = detail_index or {}
    rows_by_base = {}
    for row in progress_rows:
        if row["基地"] == "整体" or row["渠道"] == "合计":
            continue
        if row["月度目标"] == 0 and row["实际入培数"] == 0:
            continue
        rows_by_base.setdefault(row["基地"], []).append(row)

    reasons = []
    for base in [row["基地"] for row in base_risks if row["GAP"] < 0]:
        channels = sorted(rows_by_base.get(base, []), key=channel_sort_key)[:3]
        for channel in channels:
            channel["明细"] = detail_index.get((base, channel["渠道"]), [])
        has_structure_gap = any(row["GAP"] < 0 and parse_percent(row.get("占比GAP")) < 0 for row in channels)
        if has_structure_gap:
            judgment = "渠道结构偏差 + 整体量不足"
        else:
            judgment = "整体量不足"
        reasons.append({"基地": base, "判断": judgment, "主要缺口渠道": channels})
    return reasons


def build_channel_mix(progress_rows):
    rows = [
        row
        for row in progress_rows
        if row["基地"] == "整体" and row["渠道"] != "合计"
    ]
    return sorted(rows, key=lambda row: row["实际入培数"], reverse=True)


def build_base_channel_progress(progress_rows):
    base_total_gap = {
        row["基地"]: row["GAP"]
        for row in progress_rows
        if row["渠道"] == "合计"
    }
    rows = []
    for row in progress_rows:
        if row["月度目标"] == 0 and row["实际入培数"] == 0:
            continue
        item = dict(row)
        rate_value = parse_percent(row.get("达成率"))
        item["达成率_value"] = rate_value
        item["达成率未达成"] = rate_value < 100
        item["基地未达成"] = base_total_gap.get(row["基地"], 0) < 0
        item["渠道未达成"] = item["达成率未达成"]
        rows.append(item)
    return rows


def map_funnel_base(job_name):
    text = safe_text(job_name)
    if "江苏基地" in text and "淮安" in text:
        return "江苏基地-淮安"
    if "江苏基地" in text and "南京" in text:
        return "江苏基地-南京"
    if "湖南基地" in text and "空港" in text:
        return "湖南基地-空港"
    if "湖南基地" in text and "荷花" in text:
        return "湖南基地-荷花"
    if "联通河北" in text:
        return "联通河北"
    if "联通天津" in text:
        return "联通天津"
    if "长春基地" in text or "长春" in text:
        return "长春热线项目"
    if "济南基地" in text and "济阳" in text:
        return "济南基地-济阳"

    match = re.match(r"【([^】]+)】", text)
    if match:
        return match.group(1).strip()
    return "未匹配"


def ratio_percent(numerator, denominator):
    return format_percent(0 if denominator == 0 else numerator / denominator * 100)


def build_funnel_attribution(base_risks, funnel_path=None, year=None, month=None):
    empty = {
        "rules": {
            "目标面通率": format_percent(TARGET_INTERVIEW_PASS_RATE * 100),
            "目标面通到参培转化率": format_percent(TARGET_PASS_TO_TRAIN_RATE * 100),
            "目标到面到参培综合转化率": format_percent(TARGET_INTERVIEW_TO_TRAIN_RATE * 100),
            "目标口径": "截止目标",
        },
        "base_rows": [],
        "unmatched_jobs": [],
    }
    if not funnel_path:
        return empty

    df = pd.read_excel(funnel_path, sheet_name="表格", dtype=object)
    df = df.apply(lambda col: col.map(lambda value: value.strip() if isinstance(value, str) else value))
    if year is not None and month is not None and "面试官填写反馈时间" in df.columns:
        feedback_dates = pd.to_datetime(df["面试官填写反馈时间"], errors="coerce")
        df = df[(feedback_dates.dt.year == year) & (feedback_dates.dt.month == month)].copy()
    df["映射基地"] = df["职位名称"].map(map_funnel_base)
    df["面通"] = df["面试官反馈结果"].map(lambda value: safe_text(value) in PASS_FEEDBACK_RESULTS)

    grouped = {}
    for base, rows in df.groupby("映射基地"):
        grouped[base] = {
            "实际到面人数": int(len(rows)),
            "实际面通人数": int(rows["面通"].sum()),
            "推荐": int((rows["面试官反馈结果"] == "推荐").sum()),
            "强烈推荐": int((rows["面试官反馈结果"] == "强烈推荐").sum()),
            "不推荐": int((rows["面试官反馈结果"] == "不推荐").sum()),
            "备选": int((rows["面试官反馈结果"] == "备选").sum()),
            "岗位数": int(rows["职位名称"].nunique()),
        }

    base_rows = []
    for base_row in [row for row in base_risks if row["GAP"] < 0]:
        base = base_row["基地"]
        target = base_row["截止目标"]
        actual_train = base_row["实际入培数"]
        required_interviews = math.ceil(target / TARGET_INTERVIEW_TO_TRAIN_RATE) if target > 0 else 0
        required_pass = math.ceil(target / TARGET_PASS_TO_TRAIN_RATE) if target > 0 else 0
        funnel = grouped.get(base, {})
        actual_interviews = funnel.get("实际到面人数", 0)
        actual_pass = funnel.get("实际面通人数", 0)
        pass_rate_value = actual_pass / actual_interviews if actual_interviews else 0
        pass_to_train_value = actual_train / actual_pass if actual_pass else 0

        if target <= 0:
            judgment = "截止目标未到期"
        elif actual_train >= target:
            judgment = "截止目标已达成"
        elif actual_interviews == 0:
            judgment = "漏斗数据缺失"
        elif actual_interviews < required_interviews:
            judgment = "到面人数不足"
        elif pass_rate_value < TARGET_INTERVIEW_PASS_RATE:
            judgment = "面通率偏低"
        elif pass_to_train_value < TARGET_PASS_TO_TRAIN_RATE:
            judgment = "面通转化差"
        else:
            judgment = "其他原因"

        base_rows.append(
            {
                "基地": base,
                "招聘目标": target,
                "实际入培数": actual_train,
                "所需到面人数": required_interviews,
                "实际到面人数": actual_interviews,
                "所需面通人数": required_pass,
                "实际面通人数": actual_pass,
                "实际面通率": ratio_percent(actual_pass, actual_interviews),
                "实际面通到参培转化率": ratio_percent(actual_train, actual_pass),
                "实际到面到参培综合转化率": ratio_percent(actual_train, actual_interviews),
                "推荐": funnel.get("推荐", 0),
                "强烈推荐": funnel.get("强烈推荐", 0),
                "不推荐": funnel.get("不推荐", 0),
                "备选": funnel.get("备选", 0),
                "岗位数": funnel.get("岗位数", 0),
                "归因判断": judgment,
            }
        )

    unmatched_jobs = []
    if "未匹配" in set(df["映射基地"]):
        counts = df[df["映射基地"] == "未匹配"]["职位名称"].value_counts()
        unmatched_jobs = [
            {"职位名称": safe_text(job_name), "记录数": int(count)}
            for job_name, count in counts.items()
        ]

    empty["base_rows"] = base_rows
    empty["unmatched_jobs"] = unmatched_jobs
    return empty


def build_dashboard_data(report_path, year, month, active_path=None, leave_path=None, funnel_path=None):
    workbook = load_workbook(report_path, data_only=True)
    progress_rows = read_progress_rows(workbook)
    efficiency_summary, recruiter_details = read_efficiency(workbook)
    detail_index, recruiter_detail_index = build_detail_index(active_path, leave_path, year, month)
    for row in recruiter_details:
        row["明细"] = recruiter_detail_index.get(safe_text(row.get("姓名")), [])

    overall = next(row for row in progress_rows if row["基地"] == "整体" and row["渠道"] == "合计")
    auto_social = next(row for row in progress_rows if row["基地"] == "整体" and row["渠道"] == "自主社招")
    base_risks = build_base_risks(progress_rows)
    overall_efficiency = next((row for row in efficiency_summary if row.get("统计维度") == "整体"), {})

    actual_total = overall["实际入培数"]
    auto_social_share = 0 if actual_total == 0 else auto_social["实际入培数"] / actual_total * 100
    overview = {
        "统计月份": f"{year}-{month:02d}",
        "月度目标": overall["月度目标"],
        "截止目标": overall["截止目标"],
        "实际入培数": actual_total,
        "GAP": overall["GAP"],
        "达成率": overall["达成率"],
        "未达成基地数": sum(1 for row in base_risks if row["GAP"] < 0),
        "自主社招整体人效": safe_text(overall_efficiency.get("参培人效")),
        "自主社招占比": format_percent(auto_social_share),
    }

    return {
        "overview": overview,
        "channel_mix": build_channel_mix(progress_rows),
        "base_channel_progress": build_base_channel_progress(progress_rows),
        "base_risks": base_risks,
        "unmet_reasons": build_unmet_reasons(progress_rows, base_risks, detail_index),
        "funnel_attribution": build_funnel_attribution(base_risks, funnel_path, year, month),
        "efficiency_summary": efficiency_summary,
        "recruiter_details": recruiter_details,
    }


def metric_card(label, value, negative=False):
    tone = " danger" if negative else ""
    return f'<div class="metric{tone}"><span>{escape(label)}</span><strong>{escape(str(value))}</strong></div>'


def dashboard_filename(month):
    return f"招聘负责人看板-{month}月.html"


def parse_report_month(path):
    match = re.search(r"月度招聘达成进度-(\d+)月\.xlsx$", path.name)
    return int(match.group(1)) if match else None


def find_available_months(report_path):
    months = []
    for path in sorted(report_path.parent.glob("月度招聘达成进度-*月.xlsx")):
        month = parse_report_month(path)
        if month is not None:
            months.append(month)
    return months


def render_filter_bar(overview, available_months=None):
    year, month = overview["统计月份"].split("-", 1)
    current_month = int(month)
    available_months = sorted(set(available_months or [current_month]))
    month_options = "\n".join(
        f'<option value="{item}" data-url="{dashboard_filename(item)}"{" selected" if item == current_month else ""}>{item:02d}月</option>'
        for item in available_months
    )
    disabled = " disabled" if len(available_months) <= 1 else ""
    note = (
        f"当前看板可切换 {len(available_months)} 个月份。"
        if len(available_months) > 1
        else f"当前看板只包含 {year}-{month} 数据，暂无其他月份/年度可切换。"
    )
    return f"""
    <section class="filters">
      <label>统计口径
        <select id="scopeFilter" disabled>
          <option value="month" selected>月度</option>
          <option value="year">年度</option>
        </select>
      </label>
      <label>年份
        <select id="yearFilter" disabled>
          <option value="{escape(year)}" selected>{escape(year)}</option>
        </select>
      </label>
      <label id="monthFilterWrap">月份
        <select id="monthFilter" onchange="switchMonth(this.value)"{disabled}>
          {month_options}
        </select>
      </label>
      <span class="filter-note">{escape(note)}</span>
    </section>
    """


def render_channel_mix(rows):
    items = []
    for row in rows:
        items.append(
            f"""
            <div class="mix-card">
              <strong>{escape(row["渠道"])}</strong>
              <span>目标 {row["月度目标"]} · 截止目标 {row["截止目标"]} · 实际 {row["实际入培数"]}</span>
              <div class="mix-line">
                <em>目标占比 {escape(str(row["渠道目标占比"]))}</em>
                <em>达成占比 {escape(str(row["渠道达成占比"]))}</em>
                <em>占比GAP {escape(str(row["占比GAP"]))}</em>
              </div>
            </div>
            """
        )
    return "\n".join(items)


def render_base_risks(rows):
    items = []
    max_target = max([row["月度目标"] for row in rows] + [1])
    for row in rows:
        width = max(4, row["月度目标"] / max_target * 100)
        risk_class = "risk-danger" if row["GAP"] < 0 else "risk-ok"
        items.append(
            f"""
            <tr class="{risk_class}">
              <td>{escape(row["基地"])}</td>
              <td>{row["月度目标"]}</td>
              <td>{row["截止目标"]}</td>
              <td>{row["实际入培数"]}</td>
              <td>{row["GAP"]}</td>
              <td>{escape(str(row["达成率"]))}</td>
              <td><div class="bar"><i style="width:{width:.1f}%"></i></div></td>
            </tr>
            """
        )
    return "\n".join(items)


def render_base_channel_progress(rows):
    if not rows:
        return '<tr><td colspan="10" class="muted">暂无基地渠道达成数据</td></tr>'
    rendered = []
    groups = []
    for row in rows:
        if not groups or groups[-1][0] != row["基地"]:
            groups.append((row["基地"], []))
        groups[-1][1].append(row)
    for base, group_rows in groups:
        display_base = "整体达成" if base == "整体" else base
        for index, row in enumerate(group_rows):
            base_cell = ""
            if index == 0:
                base_cell = (
                    f'<td class="base-group-cell" rowspan="{len(group_rows)}">'
                    f"{escape(display_base)}</td>"
                )
            rate_class = "rate-danger" if row.get("达成率未达成") else ""
            rendered.append(
                f"""
                <tr class="base-group-row">
                  {base_cell}
                  <td>{escape(row["渠道"])}</td>
                  <td>{row["月度目标"]}</td>
                  <td>{row["截止目标"]}</td>
                  <td>{row["实际入培数"]}</td>
                  <td>{row["GAP"]}</td>
                  <td class="{rate_class}">{escape(str(row["达成率"]))}</td>
                  <td>{escape(str(row["渠道目标占比"]))}</td>
                  <td>{escape(str(row["渠道达成占比"]))}</td>
                  <td>{escape(str(row["占比GAP"]))}</td>
                </tr>
                """
            )
    return "\n".join(rendered)


def add_employee_payload(modal_payload, key, title, rows):
    modal_payload[key] = {
        "title": title,
        "columns": ["所属基地", "工号", "姓名", "手机号", "招聘渠道", "渠道名称", "入培日期", "在职状态", "离职日期", "在职天数"],
        "rows": rows,
    }


def add_efficiency_payload(modal_payload, key, title, rows):
    modal_payload[key] = {
        "title": title,
        "columns": ["姓名", "员工阶段", "截止参培达成", "截止参培达成率", "截止7天参培达成", "截止7天参培达成率"],
        "rows": rows,
    }


def add_recruiter_payload(modal_payload, key, title, rows):
    add_employee_payload(modal_payload, key, title, rows)


def detail_number(key, value):
    safe_key = str(key).replace("\\", "\\\\").replace("'", "\\'")
    return (
        f'<button type="button" class="detail-number" '
        f'onclick="openDetail(\'{escape(safe_key)}\')">'
        f"{escape(str(value))}</button>"
    )


def render_unmet_reasons(reasons, modal_payload):
    sections = []
    for reason_idx, reason in enumerate(reasons):
        channels = []
        for channel_idx, row in enumerate(reason["主要缺口渠道"]):
            detail_rows = row.get("明细", [])
            detail_key = f"unmet-{reason_idx}-{channel_idx}"
            add_employee_payload(
                modal_payload,
                detail_key,
                f'{reason["基地"]} - {row["渠道"]} 推荐明细',
                detail_rows,
            )
            channels.append(
                f"""
                <tr>
                  <td>{escape(row["渠道"])}</td>
                  <td>{row["月度目标"]}</td>
                  <td>{row["截止目标"]}</td>
                  <td>{detail_number(detail_key, row["实际入培数"])}</td>
                  <td class="negative">{row["GAP"]}</td>
                  <td>{escape(str(row["达成率"]))}</td>
                  <td>{escape(str(row["渠道目标占比"]))}</td>
                  <td>{escape(str(row["渠道达成占比"]))}</td>
                  <td>{escape(str(row["占比GAP"]))}</td>
                </tr>
                """
            )
        sections.append(
            f"""
            <article class="reason-card">
              <div>
                <h3>{escape(reason["基地"])}</h3>
                <p>{escape(reason["判断"])}</p>
              </div>
              <table>
                <thead>
                  <tr>
                    <th>渠道</th><th>月度目标</th><th>截止目标</th><th>实际入培</th>
                    <th>GAP</th><th>达成率</th><th>目标占比</th><th>达成占比</th><th>占比GAP</th>
                  </tr>
                </thead>
                <tbody>{"".join(channels)}</tbody>
              </table>
            </article>
            """
        )
    return "\n".join(sections)


def render_efficiency_summary(rows, recruiter_details, modal_payload):
    return "\n".join(
        f"""
        <tr>
          <td>{escape(str(row.get("统计维度", "")))}</td>
          <td>{row.get("招聘规模", "")}</td>
          <td>{row.get("截止参培达成", "")}</td>
          <td>{escape(str(row.get("参培人效", "")))}</td>
          <td>{row.get("截止7天参培达成", "")}</td>
          <td>{escape(str(row.get("7天人效", "")))}</td>
        </tr>
        """
        for row in rows
    )


def render_recruiter_details(rows, modal_payload=None):
    rendered = []
    for idx, row in enumerate(rows):
        detail_key = f"recruiter-{idx}"
        if modal_payload is not None:
            add_recruiter_payload(
                modal_payload,
                detail_key,
                f'{safe_text(row.get("姓名"))} - 自主社招推荐明细',
                row.get("明细", []),
            )
            train_done = detail_number(detail_key, row.get("截止参培达成", ""))
            seven_done = detail_number(detail_key, row.get("截止7天参培达成", ""))
        else:
            train_done = escape(str(row.get("截止参培达成", "")))
            seven_done = escape(str(row.get("截止7天参培达成", "")))
        rendered.append(
        f"""
        <tr>
          <td>{escape(str(row.get("姓名", "")))}</td>
          <td>{escape(str(row.get("伽睿工号", "")))}</td>
          <td>{escape(str(row.get("员工阶段", "")))}</td>
          <td>{escape(str(row.get("入职日期", "")))}</td>
          <td>{row.get("月度参培目标", "")}</td>
          <td>{row.get("截止参培目标", "")}</td>
          <td>{train_done}</td>
          <td>{escape(str(row.get("截止参培达成率", "")))}</td>
          <td>{row.get("月度7天参培目标", "")}</td>
          <td>{row.get("截止7天参培目标", "")}</td>
          <td>{seven_done}</td>
          <td>{escape(str(row.get("截止7天参培达成率", "")))}</td>
        </tr>
        """
    )
    return "\n".join(rendered)


def render_tab_nav():
    tabs = [
        ("overview", "达成总览"),
        ("risk", "基地风险"),
        ("funnel", "招聘漏斗归因"),
        ("efficiency", "自主社招人效"),
    ]
    return "\n".join(
        f'<button type="button" class="tab-button{" active" if index == 0 else ""}" '
        f'data-tab-target="{key}" onclick="switchDashboardTab(\'{key}\')">{label}</button>'
        for index, (key, label) in enumerate(tabs)
    )


def render_funnel_attribution(attribution):
    rules = attribution.get("rules", {})
    base_rows = attribution.get("base_rows", [])
    unmatched_jobs = attribution.get("unmatched_jobs", [])
    rule_text = (
        f'目标面通率 {escape(str(rules.get("目标面通率", "70.00%")))} · '
        f'目标面通到参培转化率 {escape(str(rules.get("目标面通到参培转化率", "40.00%")))} · '
        f'目标到面到参培综合转化率 {escape(str(rules.get("目标到面到参培综合转化率", "28.00%")))} · '
        f'目标口径 {escape(str(rules.get("目标口径", "截止目标")))}'
    )
    if not base_rows:
        rows_html = '<tr><td colspan="12" class="muted">暂无漏斗归因数据</td></tr>'
    else:
        rows_html = "\n".join(
            f"""
            <tr>
              <td>{escape(row["基地"])}</td>
              <td>{row["招聘目标"]}</td>
              <td>{row["实际入培数"]}</td>
              <td>{row["所需到面人数"]}</td>
              <td>{row["实际到面人数"]}</td>
              <td>{row["所需面通人数"]}</td>
              <td>{row["实际面通人数"]}</td>
              <td>{escape(str(row["实际面通率"]))}</td>
              <td>{escape(str(row["实际面通到参培转化率"]))}</td>
              <td>{escape(str(row["实际到面到参培综合转化率"]))}</td>
              <td>{row["岗位数"]}</td>
              <td class="negative">{escape(row["归因判断"])}</td>
            </tr>
            """
            for row in base_rows
        )
    unmatched_html = ""
    if unmatched_jobs:
        unmatched_rows = "\n".join(
            f"<tr><td>{escape(row['职位名称'])}</td><td>{row['记录数']}</td></tr>"
            for row in unmatched_jobs[:10]
        )
        unmatched_html = f"""
        <h3>未匹配岗位</h3>
        <p class="section-note">以下岗位未能映射到现有基地，暂不参与基地归因。</p>
        <table>
          <thead><tr><th>职位名称</th><th>记录数</th></tr></thead>
          <tbody>{unmatched_rows}</tbody>
        </table>
        """
    return f"""
    <p class="section-note">{rule_text}</p>
    <table>
      <thead>
        <tr>
          <th>基地</th><th>招聘目标</th><th>实际入培</th><th>所需到面</th><th>实际到面</th>
          <th>所需面通</th><th>实际面通</th><th>实际面通率</th><th>实际面通到参培转化率</th>
          <th>实际到面到参培综合转化率</th><th>岗位数</th><th>归因判断</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
    {unmatched_html}
    """


def render_dashboard_html(data, source_path, available_months=None):
    overview = data["overview"]
    modal_payload = {}
    unmet_reasons_html = render_unmet_reasons(data["unmet_reasons"], modal_payload)
    funnel_attribution_html = render_funnel_attribution(data.get("funnel_attribution", {}))
    efficiency_summary_html = render_efficiency_summary(data["efficiency_summary"], data["recruiter_details"], modal_payload)
    recruiter_details_html = render_recruiter_details(data["recruiter_details"], modal_payload)
    modal_payload_json = json.dumps(modal_payload, ensure_ascii=False).replace("</", "<\\/")
    metrics = [
        metric_card("整体目标", overview["月度目标"]),
        metric_card("截止目标", overview["截止目标"]),
        metric_card("实际入培数", overview["实际入培数"]),
        metric_card("GAP", overview["GAP"], overview["GAP"] < 0),
        metric_card("达成率", overview["达成率"], parse_percent(overview["达成率"]) < 100),
        metric_card("未达成基地数", overview["未达成基地数"], overview["未达成基地数"] > 0),
        metric_card("自主社招整体人效", overview["自主社招整体人效"]),
        metric_card("自主社招占比", overview["自主社招占比"]),
    ]
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>招聘负责人看板 - {escape(overview["统计月份"])}</title>
  <style>
    :root {{
      --bg: #f6f7f9;
      --panel: #ffffff;
      --text: #172033;
      --muted: #687385;
      --line: #d9dee8;
      --danger: #c0392b;
      --danger-bg: #fff0ee;
      --ok: #177245;
      --accent: #1f5eff;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: var(--bg);
      color: var(--text);
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Arial, "Microsoft YaHei", sans-serif;
    }}
    main {{ max-width: 1280px; margin: 0 auto; padding: 28px; }}
    header {{ display: flex; justify-content: space-between; gap: 24px; align-items: flex-end; margin-bottom: 24px; }}
    h1 {{ margin: 0 0 8px; font-size: 24px; }}
    h2 {{ margin: 0 0 14px; font-size: 18px; }}
    h3 {{ margin: 0 0 6px; font-size: 15px; }}
    p {{ margin: 0; color: var(--muted); }}
    .source {{ font-size: 12px; color: var(--muted); text-align: right; }}
    .filters {{ display: flex; gap: 12px; align-items: center; padding: 12px; margin-bottom: 14px; }}
    .filters label {{ color: var(--muted); font-size: 12px; }}
    .filter-note {{ color: var(--muted); font-size: 12px; }}
    select {{ margin-left: 6px; border: 1px solid var(--line); background: var(--panel); padding: 6px 8px; color: var(--text); }}
    .tabs {{ display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 18px; }}
    .tab-button {{ border: 1px solid var(--line); background: var(--panel); color: var(--muted); padding: 10px 14px; cursor: pointer; }}
    .tab-button.active {{ color: var(--accent); border-color: var(--accent); font-weight: 700; }}
    .tab-panel {{ display: none; }}
    .tab-panel.active {{ display: block; }}
    .metrics {{ display: grid; grid-template-columns: repeat(8, minmax(0, 1fr)); gap: 12px; margin-bottom: 18px; }}
    .metric {{ background: var(--panel); border: 1px solid var(--line); padding: 16px; }}
    .metric span {{ display: block; color: var(--muted); font-size: 12px; margin-bottom: 10px; }}
    .metric strong {{ font-size: 22px; }}
    .metric.danger strong {{ color: var(--danger); }}
    section {{ background: var(--panel); border: 1px solid var(--line); padding: 18px; margin-bottom: 18px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th, td {{ border-bottom: 1px solid var(--line); padding: 9px 8px; text-align: left; }}
    th {{ color: var(--muted); font-weight: 600; }}
    tr.risk-danger td:first-child, .negative {{ color: var(--danger); font-weight: 700; }}
    tr.risk-ok td:first-child {{ color: var(--ok); font-weight: 700; }}
    .base-group-cell {{ background: #f7f9fd; border-right: 2px solid var(--accent); font-weight: 700; vertical-align: middle; }}
    .base-group-row td {{ border-bottom-color: #e8ecf3; }}
    .rate-danger {{ color: var(--danger); font-weight: 700; }}
    .bar {{ width: 100%; height: 8px; background: #eef1f6; }}
    .bar i {{ display: block; height: 8px; background: var(--accent); }}
    .mix-grid {{ display: grid; grid-template-columns: repeat(5, minmax(0, 1fr)); gap: 10px; }}
    .mix-card {{ border: 1px solid var(--line); padding: 12px; background: #fbfcfe; }}
    .mix-card strong, .mix-card span, .mix-card em {{ display: block; }}
    .mix-card span, .mix-card em {{ color: var(--muted); font-size: 12px; font-style: normal; margin-top: 6px; }}
    .mix-line {{ margin-top: 8px; }}
    .section-note {{ margin: 0 0 12px; font-size: 12px; color: var(--muted); }}
    .reason-list {{ display: grid; gap: 14px; }}
    .reason-card {{ border: 1px solid var(--line); padding: 14px; background: #fbfcfe; }}
    .reason-card > div {{ display: flex; justify-content: space-between; gap: 16px; align-items: baseline; margin-bottom: 10px; }}
    .reason-card p {{ color: var(--danger); font-weight: 700; }}
    .detail-number {{ border: 0; background: transparent; color: var(--accent); font-weight: 700; padding: 0; cursor: pointer; text-decoration: underline; text-underline-offset: 3px; }}
    .detail-number:hover {{ color: var(--danger); }}
    .modal-mask {{ position: fixed; inset: 0; display: none; background: rgba(23, 32, 51, 0.45); z-index: 20; padding: 28px; }}
    .modal-mask.open {{ display: block; }}
    .modal-panel {{ width: min(1180px, 100%); max-height: calc(100vh - 56px); margin: 0 auto; background: var(--panel); border: 1px solid var(--line); display: flex; flex-direction: column; }}
    .modal-head {{ display: flex; justify-content: space-between; gap: 16px; align-items: center; border-bottom: 1px solid var(--line); padding: 16px 18px; }}
    .modal-head h2 {{ margin: 0; }}
    .modal-close {{ border: 1px solid var(--line); background: var(--panel); padding: 6px 10px; cursor: pointer; }}
    .modal-body {{ overflow: auto; padding: 14px 18px 18px; }}
    .column-filter {{ width: 100%; border: 1px solid var(--line); padding: 5px 6px; font-size: 12px; }}
    .muted {{ color: var(--muted); text-align: center; }}
    @media (max-width: 980px) {{
      main {{ padding: 18px; }}
      header {{ display: block; }}
      .metrics {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
      .mix-grid {{ grid-template-columns: 1fr; }}
      .source {{ text-align: left; margin-top: 10px; }}
    }}
  </style>
</head>
<body>
  <main>
    <header>
      <div>
        <h1>招聘负责人看板</h1>
        <p>{escape(overview["统计月份"])} · 展示全局基地风险、未达成原因与自主社招人效</p>
      </div>
      <div class="source">来源：{escape(str(source_path))}</div>
    </header>

    {render_filter_bar(overview, available_months)}

    <nav class="tabs" aria-label="看板页签">
      {render_tab_nav()}
    </nav>

    <div class="tab-panel active" data-tab-panel="overview">
      <div class="metrics">{"".join(metrics)}</div>

      <section>
        <h2>全局渠道占比</h2>
        <div class="mix-grid">{render_channel_mix(data["channel_mix"])}</div>
      </section>

      <section>
        <h2>各基地渠道达成明细</h2>
        <p class="section-note">标红表示基地或渠道当前未达成。</p>
        <table>
          <thead>
            <tr>
              <th>基地</th><th>渠道</th><th>月度目标</th><th>截止目标</th><th>实际入培</th>
              <th>GAP</th><th>达成率</th><th>目标占比</th><th>达成占比</th><th>占比GAP</th>
            </tr>
          </thead>
          <tbody>{render_base_channel_progress(data.get("base_channel_progress", []))}</tbody>
        </table>
      </section>
    </div>

    <div class="tab-panel" data-tab-panel="risk">
      <section>
        <h2>基地风险</h2>
        <table>
          <thead>
            <tr><th>基地</th><th>月度目标</th><th>截止目标</th><th>实际入培</th><th>GAP</th><th>达成率</th><th>目标量级</th></tr>
          </thead>
          <tbody>{render_base_risks(data["base_risks"])}</tbody>
        </table>
      </section>

      <section>
        <h2>未达成基地</h2>
        <div class="reason-list">{unmet_reasons_html}</div>
      </section>
    </div>

    <div class="tab-panel" data-tab-panel="funnel">
      <section>
        <h2>招聘漏斗归因</h2>
        {funnel_attribution_html}
      </section>
    </div>

    <div class="tab-panel" data-tab-panel="efficiency">
      <section>
        <h2>自主社招人效汇总</h2>
        <table>
          <thead>
            <tr><th>统计维度</th><th>招聘规模</th><th>参培达成</th><th>参培人效</th><th>7天达成</th><th>7天人效</th></tr>
          </thead>
          <tbody>{efficiency_summary_html}</tbody>
        </table>
      </section>

      <section>
        <h2>自主社招人员明细</h2>
        <table>
          <thead>
            <tr>
              <th>招聘专员姓名</th><th>伽睿工号</th><th>员工阶段</th><th>入职日期</th>
              <th>月度参培目标</th><th>截止参培目标</th><th>截止参培达成</th><th>截止参培达成率</th>
              <th>月度7天参培目标</th><th>截止7天参培目标</th><th>截止7天参培达成</th><th>截止7天参培达成率</th>
            </tr>
          </thead>
          <tbody>{recruiter_details_html}</tbody>
        </table>
      </section>
    </div>
  </main>
  <div id="detailModal" class="modal-mask" onclick="closeDetail(event)">
    <div class="modal-panel" role="dialog" aria-modal="true" aria-labelledby="detailModalTitle" onclick="event.stopPropagation()">
      <div class="modal-head">
        <h2 id="detailModalTitle">明细</h2>
        <button type="button" class="modal-close" onclick="closeDetail()">关闭</button>
      </div>
      <div id="detailModalBody" class="modal-body"></div>
    </div>
  </div>
  <script id="detailPayload" type="application/json">{modal_payload_json}</script>
  <script>
    const detailPayload = JSON.parse(document.getElementById('detailPayload').textContent);
    const scopeFilter = document.getElementById('scopeFilter');
    const monthFilterWrap = document.getElementById('monthFilterWrap');
    scopeFilter.addEventListener('change', () => {{
      monthFilterWrap.style.display = scopeFilter.value === 'year' ? 'none' : 'inline-block';
    }});
    function switchMonth(month) {{
      const option = document.querySelector(`#monthFilter option[value="${{month}}"]`);
      window.location.href = option.dataset.url;
    }}
    function switchDashboardTab(tab) {{
      document.querySelectorAll('[data-tab-target]').forEach((button) => {{
        button.classList.toggle('active', button.dataset.tabTarget === tab);
      }});
      document.querySelectorAll('[data-tab-panel]').forEach((panel) => {{
        panel.classList.toggle('active', panel.dataset.tabPanel === tab);
      }});
    }}
    function escapeHtml(value) {{
      return String(value ?? '').replace(/[&<>"']/g, (char) => ({{
        '&': '&amp;',
        '<': '&lt;',
        '>': '&gt;',
        '"': '&quot;',
        "'": '&#39;'
      }}[char]));
    }}
    function openDetail(key) {{
      const payload = detailPayload[key];
      const modal = document.getElementById('detailModal');
      const title = document.getElementById('detailModalTitle');
      const body = document.getElementById('detailModalBody');
      title.textContent = payload.title;
      if (!payload.rows.length) {{
        body.innerHTML = '<p class="muted">暂无匹配人员明细</p>';
      }} else {{
        const head = payload.columns.map((col) => `<th>${{escapeHtml(col)}}</th>`).join('');
        const filters = payload.columns.map((col, index) => (
          `<th><input class="column-filter" data-filter-col="${{index}}" placeholder="筛选${{escapeHtml(col)}}"></th>`
        )).join('');
        const rows = payload.rows.map((row) => {{
          const cells = payload.columns.map((col) => `<td>${{escapeHtml(row[col])}}</td>`).join('');
          return `<tr>${{cells}}</tr>`;
        }}).join('');
        body.innerHTML = `<table id="modalDetailTable"><thead><tr>${{head}}</tr><tr>${{filters}}</tr></thead><tbody>${{rows}}</tbody></table>`;
        body.querySelectorAll('[data-filter-col]').forEach((input) => {{
          input.addEventListener('input', filterModalTable);
        }});
      }}
      modal.classList.add('open');
    }}
    function filterModalTable(event) {{
      const table = document.getElementById('modalDetailTable');
      const filters = Array.from(table.querySelectorAll('[data-filter-col]')).map((input) => ({{
        column: Number(input.dataset.filterCol),
        value: input.value.trim().toLowerCase()
      }}));
      table.querySelectorAll('tbody tr').forEach((row) => {{
        const cells = Array.from(row.children);
        const visible = filters.every((filter) => {{
          if (!filter.value) return true;
          return cells[filter.column].textContent.toLowerCase().includes(filter.value);
        }});
        row.style.display = visible ? '' : 'none';
      }});
    }}
    function closeDetail(event) {{
      if (event && event.target.id !== 'detailModal') return;
      document.getElementById('detailModal').classList.remove('open');
    }}
    document.addEventListener('keydown', (event) => {{
      if (event.key === 'Escape') closeDetail();
    }});
  </script>
</body>
</html>
"""


def main():
    parser = argparse.ArgumentParser(description="Generate recruitment owner dashboard.")
    parser.add_argument("--report", required=True, help="月度招聘达成进度 Excel 文件")
    parser.add_argument("--active", default=None, help="在职员工信息 Excel 文件，用于展开人员明细")
    parser.add_argument("--leave", default=None, help="离职员工信息 Excel 文件，用于展开人员明细")
    parser.add_argument("--funnel", default=None, help="Moka 招聘漏斗 Excel 文件，用于漏斗归因")
    parser.add_argument("--year", type=int, required=True)
    parser.add_argument("--month", type=int, required=True)
    parser.add_argument("--output", default=None, help="输出 HTML；默认：招聘负责人看板/招聘负责人看板-X月.html")
    args = parser.parse_args()

    report_path = Path(args.report)
    output_path = Path(args.output) if args.output else Path("招聘负责人看板") / f"招聘负责人看板-{args.month}月.html"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    data = build_dashboard_data(report_path, args.year, args.month, args.active, args.leave, args.funnel)
    output_path.write_text(render_dashboard_html(data, report_path, find_available_months(report_path)), encoding="utf-8")
    print({"output": str(output_path), "unmet_bases": data["overview"]["未达成基地数"]})


if __name__ == "__main__":
    main()
