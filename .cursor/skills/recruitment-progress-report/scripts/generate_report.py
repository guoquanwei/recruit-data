#!/usr/bin/env python3
import argparse
import calendar
import shutil
from collections import OrderedDict
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


CHANNELS = ["回流", "内推", "渠道社招", "渠道校招", "自主社招"]
CHANNEL_ROWS = CHANNELS + ["合计"]
TARGET_BASE_ALIASES = {"15升投": "10015升投"}


def pct_text(num, den):
    if not den:
        return "0.00%"
    return f"{num / den * 100:.2f}%"


def share_gap_text(actual_num, actual_den, target_num, target_den):
    actual_share = 0 if not actual_den else actual_num / actual_den
    target_share = 0 if not target_den else target_num / target_den
    return f"{(actual_share - target_share) * 100:.2f}%"


def efficiency_text(done, scale):
    if not scale:
        return "0.0"
    return f"{done / scale:.1f}"


def norm_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def round_half_up(value):
    return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def fmt_date(value):
    if pd.isna(value):
        return ""
    return pd.Timestamp(value).strftime("%Y-%m-%d")


def parse_department(row, table_type):
    if table_type == "active":
        dept = norm_text(row.get("部门", ""))
        sep = "/"
    else:
        dept = norm_text(row.get("离职前部门", ""))
        sep = "-"

    office = norm_text(row.get("办公地点", ""))
    parts = [p.strip() for p in dept.split(sep) if p and p.strip()]
    ops_idx = next((i for i, p in enumerate(parts) if p.upper() == "OPS"), None)
    after = parts[ops_idx + 1 :] if ops_idx is not None else parts

    l1 = after[0] if len(after) > 0 else ""
    l2 = after[1] if len(after) > 1 else ""
    l3 = after[2] if len(after) > 2 else ""

    if l1 in {"京津冀大区", "长春基地"} and l2:
        return l2
    if l1 == "江苏基地" and l3:
        return f"{l1}-{l3[:2]}"
    if l1 == "湖南基地":
        if "荷花" in office or "荷花" in dept:
            return "湖南基地-荷花"
        if "空港" in office or "空港" in dept:
            return "湖南基地-空港"
        return "湖南基地"
    if l1 == "济南基地":
        if "济阳" in office or "济阳" in dept:
            return "济南基地-济阳"
        if "夏都" in office or "夏都" in dept:
            return "济南基地-夏都"
        return "济南基地"
    return l1


def read_target(path, cutoff_day):
    df = pd.read_excel(path, sheet_name="整体目标", dtype=object)
    df.columns = [str(c).strip() for c in df.columns]
    df["基地"] = df["基地"].map(norm_text).replace(TARGET_BASE_ALIASES)
    df["渠道"] = df["渠道"].map(norm_text)

    month_days = [f"{i}日" for i in range(1, 32) if f"{i}日" in df.columns]
    cutoff_days = [f"{i}日" for i in range(1, cutoff_day + 1) if f"{i}日" in df.columns]
    for col in month_days:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["月度目标_calc"] = df[month_days].sum(axis=1)
    df["截止目标_calc"] = df[cutoff_days].sum(axis=1)
    base_order = list(OrderedDict.fromkeys(df["基地"].tolist()))
    grouped = df.groupby(["基地", "渠道"], dropna=False)[["月度目标_calc", "截止目标_calc"]].sum()
    return grouped, base_order


def archive_target_file(target_path):
    archive_dir = Path("人才开发目标拆解归档")
    archive_dir.mkdir(parents=True, exist_ok=True)
    destination = archive_dir / target_path.name
    if target_path.resolve() != destination.resolve():
        shutil.copy2(target_path, destination)
    return destination


def read_employee(path, table_type):
    df = pd.read_excel(path, dtype=str)
    df = df.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x))
    df["来源表"] = "在职" if table_type == "active" else "离职"
    df["归类基地"] = df.apply(lambda row: parse_department(row, table_type), axis=1)
    return df


def recruiter_name(value):
    text = norm_text(value)
    if "+" in text:
        return text.split("+", 1)[0].strip()
    return text


def overlap_days(start_a, end_a, start_b, end_b):
    start = max(start_a, start_b)
    end = min(end_a, end_b)
    if pd.isna(start) or pd.isna(end) or start > end:
        return 0
    return int((end - start).days) + 1


def build_efficiency_data(employees, year, month, end_date):
    month_start = pd.Timestamp(year, month, 1)
    month_end = pd.Timestamp(year, month, calendar.monthrange(year, month)[1])
    month_days = calendar.monthrange(year, month)[1]

    emp = employees.copy()
    emp["入职日期_dt"] = pd.to_datetime(emp["入职日期"], errors="coerce")
    emp["离职日期_dt"] = pd.to_datetime(emp["离职日期"], errors="coerce")
    emp["姓名"] = emp["姓名"].map(norm_text)
    emp["工号"] = emp["工号"].map(norm_text)

    def in_month_scope(df):
        return df[
            (df["入职日期_dt"].notna())
            & (df["入职日期_dt"] <= month_end)
            & (df["离职日期_dt"].isna() | (df["离职日期_dt"] >= month_start))
        ].copy()

    active_recruiters = emp[
        (emp["来源表"] == "在职")
        & (emp["职位"].map(norm_text) == "招聘专员")
        & (emp["部门"].map(norm_text).str.contains("人才开发部", na=False))
    ].copy()
    leave_recruiters = emp[
        (emp["来源表"] == "离职")
        & (emp["离职前职位"].map(norm_text) == "招聘专员")
        & (emp["离职前部门"].map(norm_text).str.contains("人才开发部", na=False))
    ].copy()
    recruiters = pd.concat([active_recruiters, leave_recruiters], ignore_index=True, sort=False)
    recruiters = in_month_scope(recruiters)
    recruiters = recruiters.sort_values(["姓名", "工号", "入职日期_dt"]).drop_duplicates(subset=["工号"], keep="last")

    active_supervisors = emp[
        (emp["来源表"] == "在职")
        & (emp["职位"].map(norm_text) == "初级招聘主管")
        & (emp["部门"].map(norm_text).str.contains("人才开发部", na=False))
    ].copy()
    leave_supervisors = emp[
        (emp["来源表"] == "离职")
        & (emp["离职前职位"].map(norm_text) == "初级招聘主管")
        & (emp["离职前部门"].map(norm_text).str.contains("人才开发部", na=False))
    ].copy()
    supervisors = pd.concat([active_supervisors, leave_supervisors], ignore_index=True, sort=False)
    supervisors = in_month_scope(supervisors)
    supervisors = supervisors.sort_values(["姓名", "工号", "入职日期_dt"]).drop_duplicates(subset=["工号"], keep="last")

    auto = emp[emp["招聘渠道"].map(norm_text) == "自主社招"].copy()
    auto["渠道姓名"] = auto["渠道名称"].map(recruiter_name)
    auto["入培时间_dt"] = pd.to_datetime(auto["入培时间"], errors="coerce")
    auto["离职日期_dt"] = pd.to_datetime(auto["离职日期"], errors="coerce")
    auto["工号"] = auto["工号"].map(norm_text)

    train_actual = auto[
        (auto["入培时间_dt"] >= month_start)
        & (auto["入培时间_dt"] <= end_date)
        & (auto["工号"] != "")
    ].drop_duplicates(subset=["工号"])

    auto["7天发生日期"] = auto["入培时间_dt"] + pd.Timedelta(days=6)
    auto["参培结束日期"] = auto["离职日期_dt"].fillna(end_date)
    auto["参培天数"] = (auto["参培结束日期"] - auto["入培时间_dt"]).dt.days + 1
    seven_actual = auto[
        (auto["7天发生日期"] >= month_start)
        & (auto["7天发生日期"] <= end_date)
        & (auto["参培天数"] >= 7)
        & (auto["工号"] != "")
    ].drop_duplicates(subset=["工号"])

    train_counts = train_actual.groupby("渠道姓名")["工号"].nunique()
    seven_counts = seven_actual.groupby("渠道姓名")["工号"].nunique()

    rows = []
    for _, rec in recruiters.iterrows():
        hire = rec["入职日期_dt"]
        leave = rec["离职日期_dt"]
        stage = "正式期" if hire + pd.DateOffset(months=6) <= month_end else "试用期"
        train_standard = 20 if stage == "正式期" else 12
        seven_standard = 12 if stage == "正式期" else 8
        monthly_days = overlap_days(hire, leave if pd.notna(leave) else month_end, month_start, month_end)
        cutoff_days = overlap_days(hire, leave if pd.notna(leave) else end_date, month_start, end_date)
        monthly_train_target = round_half_up(train_standard * monthly_days / month_days)
        cutoff_train_target = round_half_up(train_standard * cutoff_days / month_days)
        monthly_seven_target = round_half_up(seven_standard * monthly_days / month_days)
        cutoff_seven_target = round_half_up(seven_standard * cutoff_days / month_days)
        name = rec["姓名"]
        train_done = int(train_counts.get(name, 0))
        seven_done = int(seven_counts.get(name, 0))
        rows.append(
            {
                "姓名": name,
                "伽睿工号": rec["工号"],
                "员工阶段": stage,
                "入职日期": fmt_date(hire),
                "离职日期": fmt_date(leave),
                "月度参培目标": monthly_train_target,
                "截止参培目标": cutoff_train_target,
                "截止参培达成": train_done,
                "截止参培达成率": pct_text(train_done, cutoff_train_target),
                "月度7天参培目标": monthly_seven_target,
                "截止7天参培目标": cutoff_seven_target,
                "截止7天参培达成": seven_done,
                "截止7天参培达成率": pct_text(seven_done, cutoff_seven_target),
            }
        )

    detail = pd.DataFrame(rows)
    duplicate_names = (
        recruiters.groupby("姓名")["工号"].nunique().reset_index(name="工号数").query("工号数 > 1")
        if not recruiters.empty
        else pd.DataFrame(columns=["姓名", "工号数"])
    )
    supervisor_names = set(supervisors["姓名"].dropna().map(norm_text))
    train_supervisor_index = train_counts.index.intersection(supervisor_names)
    seven_supervisor_index = seven_counts.index.intersection(supervisor_names)
    supervisor_actual = {
        "截止参培达成": int(train_counts.loc[train_supervisor_index].sum()) if len(train_supervisor_index) else 0,
        "截止7天参培达成": int(seven_counts.loc[seven_supervisor_index].sum()) if len(seven_supervisor_index) else 0,
        "人数": int(len(supervisors)),
    }
    return detail, duplicate_names, supervisor_actual


def make_group(base, channel_values):
    total_target = sum(channel_values[ch]["月度目标"] for ch in CHANNELS)
    total_cutoff = sum(channel_values[ch]["截止目标"] for ch in CHANNELS)
    total_actual = sum(channel_values[ch]["实际入培数"] for ch in CHANNELS)

    rows = []
    for ch in CHANNELS:
        monthly = channel_values[ch]["月度目标"]
        cutoff = channel_values[ch]["截止目标"]
        actual = channel_values[ch]["实际入培数"]
        rows.append(
            {
                "基地": base,
                "渠道": ch,
                "月度目标": monthly,
                "截止目标": cutoff,
                "实际入培数": actual,
                "GAP": actual - monthly,
                "达成率": pct_text(actual, monthly),
                "渠道目标占比": pct_text(monthly, total_target),
                "渠道达成占比": pct_text(actual, total_actual),
                "占比GAP": share_gap_text(actual, total_actual, monthly, total_target),
                "is_total": False,
            }
        )

    rows.append(
        {
            "基地": base,
            "渠道": "合计",
            "月度目标": total_target,
            "截止目标": total_cutoff,
            "实际入培数": total_actual,
            "GAP": total_actual - total_target,
            "达成率": pct_text(total_actual, total_target),
            "渠道目标占比": "100.00%" if total_target else "0.00%",
            "渠道达成占比": "100.00%" if total_actual else "0.00%",
            "占比GAP": "0.00%",
            "is_total": True,
        }
    )
    return rows


def build_report(target_group, base_order, actual_group):
    detail = OrderedDict()
    for base in base_order:
        detail[base] = {ch: {"月度目标": 0, "截止目标": 0, "实际入培数": 0} for ch in CHANNELS}
        for ch in CHANNELS:
            if (base, ch) in target_group.index:
                detail[base][ch]["月度目标"] = int(target_group.loc[(base, ch), "月度目标_calc"])
                detail[base][ch]["截止目标"] = int(target_group.loc[(base, ch), "截止目标_calc"])
            if (base, ch) in actual_group.index:
                detail[base][ch]["实际入培数"] = int(actual_group.loc[(base, ch)])

    overall = {ch: {"月度目标": 0, "截止目标": 0, "实际入培数": 0} for ch in CHANNELS}
    for base in base_order:
        for ch in CHANNELS:
            for key in overall[ch]:
                overall[ch][key] += detail[base][ch][key]

    rows = make_group("整体", overall)
    for base in base_order:
        rows.extend(make_group(base, detail[base]))
    return rows


def write_efficiency_sheet(wb, efficiency_detail, duplicate_names, supervisor_actual):
    ws = wb.create_sheet("人效核算", 1)
    show_leave = not efficiency_detail.empty and efficiency_detail["离职日期"].astype(str).str.strip().ne("").any()
    headers = [
        "姓名",
        "伽睿工号",
        "员工阶段",
        "入职日期",
    ]
    if show_leave:
        headers.append("离职日期")
    headers += [
        "月度参培目标",
        "截止参培目标",
        "截止参培达成",
        "截止参培达成率",
        "月度7天参培目标",
        "截止7天参培目标",
        "截止7天参培达成",
        "截止7天参培达成率",
    ]

    header_fill = PatternFill("solid", fgColor="9DC3E6")
    thin = Side(style="thin", color="808080")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for col, title in enumerate(headers, 1):
        cell = ws.cell(1, col, title)
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    sorted_detail = efficiency_detail.sort_values(["员工阶段", "姓名", "伽睿工号"]) if not efficiency_detail.empty else efficiency_detail
    for row_idx, (_, row) in enumerate(sorted_detail.iterrows(), 2):
        values = [row["姓名"], row["伽睿工号"], row["员工阶段"], row["入职日期"]]
        if show_leave:
            values.append(row["离职日期"])
        values += [
            row["月度参培目标"],
            row["截止参培目标"],
            row["截止参培达成"],
            row["截止参培达成率"],
            row["月度7天参培目标"],
            row["截止7天参培目标"],
            row["截止7天参培达成"],
            row["截止7天参培达成率"],
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col, value)
            cell.border = border
            cell.alignment = center
            if headers[col - 1].endswith("达成率"):
                try:
                    rate = float(str(value).replace("%", ""))
                except ValueError:
                    rate = 0
                if rate < 70:
                    cell.font = Font(color="C00000", bold=True)
                elif rate < 100:
                    cell.font = Font(color="C00000")

    summary_start = len(sorted_detail) + 4
    summary_headers = ["统计维度", "招聘规模", "截止参培达成", "参培人效", "截止7天参培达成", "7天人效"]
    for col, title in enumerate(summary_headers, 1):
        cell = ws.cell(summary_start, col, title)
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    summary_rows = []
    for stage in ["试用期", "正式期"]:
        stage_df = sorted_detail[sorted_detail["员工阶段"] == stage] if not sorted_detail.empty else sorted_detail
        scale = int(len(stage_df))
        train_done = int(stage_df["截止参培达成"].sum()) if scale else 0
        seven_done = int(stage_df["截止7天参培达成"].sum()) if scale else 0
        summary_rows.append([stage, scale, train_done, efficiency_text(train_done, scale), seven_done, efficiency_text(seven_done, scale)])
    scale = int(len(sorted_detail))
    train_done = (int(sorted_detail["截止参培达成"].sum()) if scale else 0) + int(supervisor_actual.get("截止参培达成", 0))
    seven_done = (int(sorted_detail["截止7天参培达成"].sum()) if scale else 0) + int(supervisor_actual.get("截止7天参培达成", 0))
    summary_rows.append(["整体", scale, train_done, efficiency_text(train_done, scale), seven_done, efficiency_text(seven_done, scale)])

    for row_idx, row in enumerate(summary_rows, summary_start + 1):
        for col, value in enumerate(row, 1):
            cell = ws.cell(row_idx, col, value)
            cell.border = border
            cell.alignment = center
            if row[0] == "整体":
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.font = Font(bold=True)

    if not duplicate_names.empty:
        note_row = summary_start + len(summary_rows) + 3
        ws.cell(note_row, 1, "重名提示")
        ws.cell(note_row, 1).font = Font(bold=True, color="C00000")
        for idx, (_, row) in enumerate(duplicate_names.iterrows(), note_row + 1):
            ws.cell(idx, 1, row["姓名"])
            ws.cell(idx, 2, int(row["工号数"]))

    widths = [12, 14, 10, 12, 12, 14, 14, 14, 16, 16, 16, 16, 18]
    for col, width in enumerate(widths[: len(headers)], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"


def write_workbook(output_path, report_rows, unmatched_summary, efficiency_detail, duplicate_names, supervisor_actual, year, month, start_date, end_date):
    wb = Workbook()
    ws = wb.active
    ws.title = "达成进度"

    headers = ["基地", "渠道", "月度目标", "截止目标", "实际入培数", "GAP", "达成率", "渠道目标占比", "渠道达成占比", "占比GAP"]
    header_row = 1
    first_data_row = 2

    header_fill = PatternFill("solid", fgColor="9DC3E6")
    for col, title in enumerate(headers, 1):
        cell = ws.cell(header_row, col, title)
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="808080")
    all_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    overall_fill = PatternFill("solid", fgColor="BDD7EE")
    overall_detail_fill = PatternFill("solid", fgColor="EAF2F8")

    for idx, row in enumerate(report_rows, first_data_row):
        for col, title in enumerate(headers, 1):
            cell = ws.cell(idx, col, row[title])
            cell.border = all_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row["is_total"]:
                cell.fill = overall_fill if row["基地"] == "整体" else total_fill
                cell.font = Font(bold=True)
            elif row["基地"] == "整体":
                cell.fill = overall_detail_fill

    current = first_data_row
    for row in report_rows[:: len(CHANNEL_ROWS)]:
        start = current
        end = current + len(CHANNEL_ROWS) - 1
        ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
        ws.cell(start, 1).value = row["基地"]
        ws.cell(start, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(start, 1).font = Font(bold=True)
        ws.cell(start, 1).border = all_border
        for merge_row in range(start, end + 1):
            ws.cell(merge_row, 1).border = all_border
        current = end + 1

    for row in ws.iter_rows(min_row=header_row, max_row=first_data_row + len(report_rows) - 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = all_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(header_row, col)
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill
        cell.border = all_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(first_data_row, first_data_row + len(report_rows)):
        base_cell = ws.cell(row, 1)
        if base_cell.value is not None:
            base_cell.font = Font(bold=True, color="000000")
            base_cell.border = all_border
            base_cell.alignment = Alignment(horizontal="center", vertical="center")
        rate_cell = ws.cell(row, headers.index("达成率") + 1)
        try:
            rate = float(str(rate_cell.value or "0").replace("%", ""))
        except ValueError:
            rate = 0
        if rate < 70:
            rate_cell.font = Font(color="C00000", bold=True)
        elif rate < 100:
            rate_cell.font = Font(color="C00000")

    widths = [22, 12, 12, 12, 12, 10, 12, 14, 14, 12]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    summary = wb.create_sheet("汇总")
    overall_total = next(row for row in report_rows if row["基地"] == "整体" and row["渠道"] == "合计")
    summary_rows = [
        ["统计月份", f"{year}-{month:02d}"],
        ["月度目标合计", overall_total["月度目标"]],
        ["实际入培合计", overall_total["实际入培数"]],
        ["GAP合计", overall_total["GAP"]],
        ["整体达成率", overall_total["达成率"]],
        ["未匹配目标基地的实际入培人数", int(unmatched_summary["实际入培数"].sum()) if not unmatched_summary.empty else 0],
    ]
    for r_idx, row in enumerate(summary_rows, 1):
        for c_idx, value in enumerate(row, 1):
            summary.cell(r_idx, c_idx, value)
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 18

    unmatched = wb.create_sheet("未匹配实际数据")
    unmatched.append(["归类基地", "招聘渠道", "实际入培数", "说明"])
    if unmatched_summary.empty:
        unmatched.append(["", "", 0, "无未匹配目标基地的实际入培数据"])
    else:
        for _, row in unmatched_summary.iterrows():
            unmatched.append([row["归类基地"], row["招聘渠道"], int(row["实际入培数"]), "该归类基地未出现在目标表基地列表中"])
    for col, width in enumerate([24, 14, 12, 44], 1):
        unmatched.column_dimensions[get_column_letter(col)].width = width

    write_efficiency_sheet(wb, efficiency_detail, duplicate_names, supervisor_actual)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Generate recruitment progress report.")
    parser.add_argument("--target", required=True, help="人才开发目标拆解 Excel 文件")
    parser.add_argument("--active", required=True, help="在职员工信息 Excel 文件")
    parser.add_argument("--leave", required=True, help="离职员工信息 Excel 文件")
    parser.add_argument("--year", type=int, required=True)
    parser.add_argument("--month", type=int, required=True)
    parser.add_argument("--output", default=None, help="输出文件；默认：月度招聘达成进度/月度招聘达成进度-X月.xlsx")
    parser.add_argument("--cutoff-day", type=int, default=None, help="截止日期日数字；默认当前月份取当天，历史月份取当月最后一天")
    args = parser.parse_args()

    last_day = calendar.monthrange(args.year, args.month)[1]
    today = pd.Timestamp.today().normalize()
    if args.cutoff_day is not None:
        cutoff_day = min(args.cutoff_day, last_day)
    elif args.year == today.year and args.month == today.month:
        cutoff_day = min(today.day, last_day)
    else:
        cutoff_day = last_day
    start_date = pd.Timestamp(args.year, args.month, 1)
    end_date = pd.Timestamp(args.year, args.month, cutoff_day)

    output_path = Path(args.output) if args.output else Path("月度招聘达成进度") / f"月度招聘达成进度-{args.month}月.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    target_path = Path(args.target)
    archived_target = archive_target_file(target_path)
    target_group, base_order = read_target(target_path, cutoff_day)

    active = read_employee(Path(args.active), "active")
    leave = read_employee(Path(args.leave), "leave")
    employees = pd.concat([active, leave], ignore_index=True, sort=False)
    employees["入培时间_dt"] = pd.to_datetime(employees["入培时间"], errors="coerce")
    employees["招聘渠道"] = employees["招聘渠道"].map(norm_text)
    employees["工号"] = employees["工号"].map(norm_text)

    actual = employees[
        (employees["入培时间_dt"] >= start_date)
        & (employees["入培时间_dt"] <= end_date)
        & (employees["招聘渠道"].isin(CHANNELS))
        & (employees["工号"] != "")
    ].copy()
    actual_unique = actual.drop_duplicates(subset=["工号"])
    actual_group = actual_unique.groupby(["归类基地", "招聘渠道"])["工号"].nunique()

    target_bases = set(base_order)
    unmatched = actual_unique[~actual_unique["归类基地"].isin(target_bases)]
    if unmatched.empty:
        unmatched_summary = pd.DataFrame(columns=["归类基地", "招聘渠道", "实际入培数"])
    else:
        unmatched_summary = unmatched.groupby(["归类基地", "招聘渠道"])["工号"].nunique().reset_index(name="实际入培数")

    efficiency_detail, duplicate_names, supervisor_actual = build_efficiency_data(employees, args.year, args.month, end_date)
    report_rows = build_report(target_group, base_order, actual_group)
    write_workbook(output_path, report_rows, unmatched_summary, efficiency_detail, duplicate_names, supervisor_actual, args.year, args.month, start_date, end_date)

    overall = next(row for row in report_rows if row["基地"] == "整体" and row["渠道"] == "合计")
    print(
        {
            "output": str(output_path.resolve()),
            "archived_target": str(archived_target.resolve()),
            "target": overall["月度目标"],
            "actual": overall["实际入培数"],
            "gap": overall["GAP"],
            "achievement_rate": overall["达成率"],
            "unmatched_actual": int(unmatched_summary["实际入培数"].sum()) if not unmatched_summary.empty else 0,
            "recruiters": int(len(efficiency_detail)),
            "junior_supervisor_actual": supervisor_actual,
            "duplicate_recruiter_names": duplicate_names.to_dict("records"),
        }
    )


if __name__ == "__main__":
    main()
