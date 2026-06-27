import importlib.util
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


SCRIPT_PATH = (
    Path(__file__).resolve().parents[1]
    / "scripts"
    / "generate_dashboard.py"
)


def load_module():
    spec = importlib.util.spec_from_file_location("generate_dashboard", SCRIPT_PATH)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


class GenerateDashboardTest(unittest.TestCase):
    def make_report(self, path):
        wb = Workbook()
        ws = wb.active
        ws.title = "达成进度"
        ws.append(
            [
                "基地",
                "渠道",
                "月度目标",
                "截止目标",
                "实际入培数",
                "GAP",
                "达成率",
                "渠道目标占比",
                "渠道达成占比",
                "占比GAP",
            ]
        )
        rows = [
            ["整体", "自主社招", 100, 80, 50, -50, "50.00%", "66.67%", "62.50%", "-4.17%"],
            [None, "内推", 50, 40, 30, -20, "60.00%", "33.33%", "37.50%", "4.17%"],
            [None, "合计", 150, 120, 80, -70, "53.33%", "100.00%", "100.00%", "0.00%"],
            ["A基地", "自主社招", 40, 30, 10, -30, "25.00%", "66.67%", "50.00%", "-16.67%"],
            [None, "内推", 10, 10, 5, -5, "50.00%", "16.67%", "25.00%", "8.33%"],
            [None, "渠道社招", 10, 10, 5, -5, "50.00%", "16.67%", "25.00%", "8.33%"],
            [None, "合计", 60, 50, 20, -40, "33.33%", "100.00%", "100.00%", "0.00%"],
            ["B基地", "自主社招", 20, 20, 30, 10, "150.00%", "100.00%", "100.00%", "0.00%"],
            [None, "合计", 20, 20, 30, 10, "150.00%", "100.00%", "100.00%", "0.00%"],
            ["空基地", "自主社招", 0, 0, 0, 0, "0.00%", "0.00%", "0.00%", "0.00%"],
            [None, "合计", 0, 0, 0, 0, "0.00%", "0.00%", "0.00%", "0.00%"],
        ]
        for row in rows:
            ws.append(row)

        eff = wb.create_sheet("人效核算")
        eff.append(
            [
                "姓名",
                "伽睿工号",
                "员工阶段",
                "入职日期",
                "月度参培目标",
                "截止参培目标",
                "截止参培达成",
                "截止参培达成率",
                "月度7天参培目标",
                "截止7天参培目标",
                "截止7天参培达成",
                "截止7天参培达成率",
            ]
        )
        eff.append(["张三", "J001", "正式期", "2025-01-01", 20, 10, 8, "80.00%", 12, 6, 5, "83.33%"])
        eff.append(["李四", "J002", "试用期", "2026-05-01", 12, 9, 3, "33.33%", 8, 6, 2, "33.33%"])
        eff.append([])
        eff.append(["统计维度", "招聘规模", "截止参培达成", "参培人效", "截止7天参培达成", "7天人效"])
        eff.append(["试用期", 1, 3, "3.0", 2, "2.0"])
        eff.append(["正式期", 1, 8, "8.0", 5, "5.0"])
        eff.append(["整体", 2, 11, "5.5", 7, "3.5"])
        wb.save(path)

    def make_employee_files(self, active_path, leave_path):
        active = Workbook()
        ws = active.active
        ws.append(
            [
                "工号",
                "姓名",
                "手机号码",
                "招聘渠道",
                "渠道名称",
                "入培时间",
                "入职日期",
                "离职日期",
                "员工状态",
                "部门",
                "办公地点",
            ]
        )
        ws.append(
            [
                "J001",
                "王五",
                "13800000001",
                "自主社招",
                "张三+J900",
                "2026/06/10",
                "2026/06/10",
                "",
                "在职",
                "伽睿集团 / OPS / A基地 / 一区",
                "A",
            ]
        )
        ws.append(
            [
                "J004",
                "周八",
                "13800000004",
                "自主社招",
                "张三+J900",
                "2026/06/13",
                "2026/06/13",
                "",
                "在职",
                "伽睿集团 / OPS / A基地 / 一区",
                "A",
            ]
        )
        ws.append(
            [
                "J002",
                "赵六",
                "13800000002",
                "内推",
                "李四+J901",
                "2026/06/11",
                "2026/06/11",
                "",
                "在职",
                "伽睿集团 / OPS / B基地 / 一区",
                "B",
            ]
        )
        active.save(active_path)

        leave = Workbook()
        ws = leave.active
        ws.append(
            [
                "工号",
                "姓名",
                "手机号码",
                "招聘渠道",
                "渠道名称",
                "入培时间",
                "入职日期",
                "离职日期",
                "离职前部门",
                "办公地点",
            ]
        )
        ws.append(
            [
                "J003",
                "孙七",
                "13800000003",
                "渠道社招",
                "外部渠道A",
                "2026/06/12",
                "2026/06/12",
                "2026/06/20",
                "伽睿集团 - OPS - A基地 - 一区",
                "A",
            ]
        )
        leave.save(leave_path)

    def make_funnel_file(self, path):
        wb = Workbook()
        ws = wb.active
        ws.title = "表格"
        ws.append(
            [
                "职位名称",
                "候选人名称",
                "性别",
                "电话",
                "面试官填写反馈时间",
                "面试官反馈结果",
                "面试官",
                "猎头公司标签",
                "猎头合约名称",
                "内推人",
                "综合评价",
            ]
        )
        for index in range(40):
            result = "推荐" if index < 20 else "不推荐"
            ws.append(
                [
                    "【A基地】热线客服",
                    f"候选人{index}",
                    "男",
                    f"1380001{index:04d}",
                    "2026-06-20",
                    result,
                    "A基地面试官",
                    "自主社招",
                    "张三+J001",
                    "-",
                    "",
                ]
            )
        wb.save(path)

    def test_build_dashboard_data_summarizes_top_metrics_and_efficiency(self):
        module = load_module()
        with tempfile.TemporaryDirectory() as tmp:
            report = Path(tmp) / "report.xlsx"
            self.make_report(report)

            data = module.build_dashboard_data(report, year=2026, month=6)

        self.assertEqual(data["overview"]["月度目标"], 150)
        self.assertEqual(data["overview"]["截止目标"], 120)
        self.assertEqual(data["overview"]["实际入培数"], 80)
        self.assertEqual(data["overview"]["GAP"], -70)
        self.assertEqual(data["overview"]["未达成基地数"], 1)
        self.assertEqual(data["overview"]["自主社招整体人效"], "5.5")
        self.assertEqual(data["overview"]["自主社招占比"], "62.50%")
        self.assertEqual(data["channel_mix"][0]["渠道"], "自主社招")
        self.assertEqual(data["channel_mix"][0]["实际入培数"], 50)

    def test_build_dashboard_data_filters_empty_bases_and_sorts_risks(self):
        module = load_module()
        with tempfile.TemporaryDirectory() as tmp:
            report = Path(tmp) / "report.xlsx"
            self.make_report(report)

            data = module.build_dashboard_data(report, year=2026, month=6)

        base_names = [row["基地"] for row in data["base_risks"]]
        self.assertEqual(base_names, ["A基地", "B基地"])
        self.assertEqual(data["base_risks"][0]["风险等级"], "未达成")
        self.assertEqual(data["base_risks"][1]["风险等级"], "已达成")

    def test_build_dashboard_data_ranks_channel_reasons_for_unmet_bases(self):
        module = load_module()
        with tempfile.TemporaryDirectory() as tmp:
            report = Path(tmp) / "report.xlsx"
            self.make_report(report)

            data = module.build_dashboard_data(report, year=2026, month=6)

        reasons = data["unmet_reasons"]
        self.assertEqual(len(reasons), 1)
        self.assertEqual(reasons[0]["基地"], "A基地")
        self.assertEqual(reasons[0]["主要缺口渠道"][0]["渠道"], "自主社招")
        self.assertEqual(reasons[0]["主要缺口渠道"][0]["GAP"], -30)
        self.assertEqual(reasons[0]["判断"], "渠道结构偏差 + 整体量不足")

    def test_build_dashboard_data_adds_base_channel_progress(self):
        module = load_module()
        with tempfile.TemporaryDirectory() as tmp:
            report = Path(tmp) / "report.xlsx"
            self.make_report(report)

            data = module.build_dashboard_data(report, year=2026, month=6)

        rows = data["base_channel_progress"]
        overall_auto = next(row for row in rows if row["基地"] == "整体" and row["渠道"] == "自主社招")
        a_auto = next(row for row in rows if row["基地"] == "A基地" and row["渠道"] == "自主社招")
        b_total = next(row for row in rows if row["基地"] == "B基地" and row["渠道"] == "合计")
        self.assertEqual(rows[0]["基地"], "整体")
        self.assertEqual(overall_auto["达成率_value"], 50.0)
        self.assertTrue(overall_auto["达成率未达成"])
        self.assertEqual(a_auto["截止目标"], 30)
        self.assertEqual(a_auto["实际入培数"], 10)
        self.assertEqual(a_auto["GAP"], -30)
        self.assertTrue(a_auto["达成率未达成"])
        self.assertFalse(b_total["达成率未达成"])

    def test_build_dashboard_data_attaches_employee_details_to_unmet_channels(self):
        module = load_module()
        with tempfile.TemporaryDirectory() as tmp:
            report = Path(tmp) / "report.xlsx"
            active = Path(tmp) / "active.xlsx"
            leave = Path(tmp) / "leave.xlsx"
            self.make_report(report)
            self.make_employee_files(active, leave)

            data = module.build_dashboard_data(
                report,
                year=2026,
                month=6,
                active_path=active,
                leave_path=leave,
            )

        channels = {
            row["渠道"]: row["明细"]
            for row in data["unmet_reasons"][0]["主要缺口渠道"]
        }
        auto_detail = channels["自主社招"]
        channel_detail = channels["渠道社招"]
        self.assertEqual(auto_detail[0]["所属基地"], "A基地")
        self.assertEqual(auto_detail[0]["工号"], "J001")
        self.assertEqual(auto_detail[0]["姓名"], "王五")
        self.assertEqual(auto_detail[0]["手机号"], "13800000001")
        self.assertEqual(auto_detail[0]["在职状态"], "在职")
        self.assertEqual(channel_detail[0]["招聘渠道"], "渠道社招")
        self.assertEqual(channel_detail[0]["在职状态"], "离职")
        self.assertEqual(channel_detail[0]["在职天数"], 9)

    def test_build_dashboard_data_adds_funnel_attribution_from_cutoff_target(self):
        module = load_module()
        with tempfile.TemporaryDirectory() as tmp:
            report = Path(tmp) / "report.xlsx"
            funnel = Path(tmp) / "funnel.xlsx"
            self.make_report(report)
            self.make_funnel_file(funnel)

            data = module.build_dashboard_data(
                report,
                year=2026,
                month=6,
                funnel_path=funnel,
            )

        rows = data["funnel_attribution"]["base_rows"]
        a_base = next(row for row in rows if row["基地"] == "A基地")
        self.assertEqual(a_base["招聘目标"], 50)
        self.assertEqual(a_base["所需到面人数"], 179)
        self.assertEqual(a_base["所需面通人数"], 125)
        self.assertEqual(a_base["实际到面人数"], 40)
        self.assertEqual(a_base["实际面通人数"], 20)
        self.assertEqual(a_base["实际面通率"], "50.00%")
        self.assertEqual(a_base["实际面通到参培转化率"], "100.00%")
        self.assertEqual(a_base["归因判断"], "到面人数不足")

    def test_render_dashboard_html_contains_modal_payload_with_real_details(self):
        module = load_module()
        with tempfile.TemporaryDirectory() as tmp:
            report = Path(tmp) / "report.xlsx"
            active = Path(tmp) / "active.xlsx"
            leave = Path(tmp) / "leave.xlsx"
            self.make_report(report)
            self.make_employee_files(active, leave)
            data = module.build_dashboard_data(
                report,
                year=2026,
                month=6,
                active_path=active,
                leave_path=leave,
            )

            html = module.render_dashboard_html(data, report)

        self.assertIn("detailModal", html)
        self.assertIn("openDetail", html)
        self.assertIn("detail-number", html)
        self.assertNotIn("查看明细", html)
        self.assertIn("filterModalTable", html)
        self.assertIn("data-filter-col", html)
        self.assertIn("<h2>未达成基地</h2>", html)
        self.assertNotIn("未达成基地（所属基地", html)
        self.assertIn("13800000001", html)
        self.assertIn("王五", html)
        self.assertIn("孙七", html)

    def test_render_dashboard_html_enables_month_switching_when_months_exist(self):
        module = load_module()
        with tempfile.TemporaryDirectory() as tmp:
            report = Path(tmp) / "report.xlsx"
            self.make_report(report)
            data = module.build_dashboard_data(report, year=2026, month=6)

            html = module.render_dashboard_html(
                data,
                report,
                available_months=[5, 6],
            )

        self.assertIn('<option value="5"', html)
        self.assertIn('<option value="6"', html)
        self.assertIn("selected>06月", html)
        self.assertIn("switchMonth", html)
        self.assertIn("招聘负责人看板-5月.html", html)
        self.assertNotIn("暂无其他月份/年度可切换", html)

    def test_render_dashboard_html_contains_tabs_and_funnel_attribution(self):
        module = load_module()
        with tempfile.TemporaryDirectory() as tmp:
            report = Path(tmp) / "report.xlsx"
            funnel = Path(tmp) / "funnel.xlsx"
            self.make_report(report)
            self.make_funnel_file(funnel)
            data = module.build_dashboard_data(
                report,
                year=2026,
                month=6,
                funnel_path=funnel,
            )

            html = module.render_dashboard_html(data, report)

        self.assertIn("data-tab-target=\"overview\"", html)
        self.assertIn("data-tab-panel=\"funnel\"", html)
        self.assertIn("招聘漏斗归因", html)
        self.assertIn("目标面通率 70.00%", html)
        self.assertIn("到面人数不足", html)
        self.assertIn("实际面通到参培转化率", html)

    def test_render_dashboard_html_contains_base_channel_progress_table(self):
        module = load_module()
        with tempfile.TemporaryDirectory() as tmp:
            report = Path(tmp) / "report.xlsx"
            self.make_report(report)
            data = module.build_dashboard_data(report, year=2026, month=6)

            html = module.render_dashboard_html(data, report)

        self.assertIn("各基地渠道达成明细", html)
        self.assertIn("整体达成", html)
        self.assertIn("base-group-cell", html)
        self.assertIn("rowspan=", html)
        self.assertIn("rate-danger", html)
        self.assertIn(">A基地</td>", html)
        self.assertIn("<td>自主社招</td>", html)

    def test_recruiter_detail_rows_have_popup_payload_and_full_efficiency_columns(self):
        module = load_module()
        with tempfile.TemporaryDirectory() as tmp:
            report = Path(tmp) / "report.xlsx"
            active = Path(tmp) / "active.xlsx"
            leave = Path(tmp) / "leave.xlsx"
            self.make_report(report)
            self.make_employee_files(active, leave)
            data = module.build_dashboard_data(
                report,
                year=2026,
                month=6,
                active_path=active,
                leave_path=leave,
            )

            html = module.render_dashboard_html(data, report)

        self.assertIn("月度参培目标", html)
        self.assertIn("截止参培目标", html)
        self.assertIn("月度7天参培目标", html)
        self.assertIn("recruiter-", html)
        self.assertIn("张三 - 自主社招推荐明细", html)
        self.assertIn("13800000004", html)
        self.assertNotIn("efficiency-0", html)


if __name__ == "__main__":
    unittest.main()
